from checkpoint import *
from openpyxl import Workbook
import os
import argparse


def get_stats(stats_file, warmup=True):
    begin_flag = "Begin Simulation Statistics"
    warmup_done = False if warmup else True
    start = False
    kv_map = {}
    with open(stats_file, 'r') as f:
        line_num = 0
        for line in f.readlines():
            line_num += 1
            if begin_flag in line:
                if warmup_done:
                    start = True
                else:
                    warmup_done = True
            elif start:
                if line == '\n':
                    break
                else:
                    lst = line.split()
                    k, v = lst[0], lst[1]
                    try:
                        int_v = int(v)
                        kv_map[k] = int_v
                    except ValueError:
                        fp_v = float(v)
                        kv_map[k] = fp_v
    return kv_map


def merge_stats(cpt_stats_lst, weight_lst):
    acc_kv_map = {}
    scale = 100
    for cpt_kv_map, cpt_weight in zip(cpt_stats_lst, weight_lst):
        for (cpt_k, cpt_v) in cpt_kv_map.items():
            cpt_weighted_v = cpt_v * scale * float(cpt_weight)
            if cpt_k not in acc_kv_map:
                acc_kv_map[cpt_k] = cpt_weighted_v
            else:
                acc_kv_map[cpt_k] += cpt_weighted_v
    for k, v in acc_kv_map.items():
        acc_kv_map[k] = v / scale
    return acc_kv_map


def merge_spec_results(cpt_dir, result_dir):
    cpts = get_checkpoints(cpt_dir)
    bmks = {}

    for checkpoint in cpts:
        tgt_dir = checkpoint.output_dir(result_dir)
        files = os.listdir(tgt_dir)
        if "completed" not in files or "stats.txt" not in files:
            print(f"Skip checkpoint: {tgt_dir}")
            continue
        if checkpoint.basename not in bmks:
            bmks[checkpoint.basename] = [checkpoint]
        else:
            bmks[checkpoint.basename].append(checkpoint)

    results = {}
    t = 0
    for benchmark, bmk_cpts in bmks.items():
        print(f"Merging {benchmark} ({len(bmk_cpts)}) checkpoints")
        t += len(bmk_cpts)
        mp = merge_stats(
            map(lambda c: get_stats(os.path.join(c.output_dir(result_dir), "stats.txt")), bmk_cpts),
            map(lambda c: c.weight, bmk_cpts)
        )
        results[benchmark] = mp
    return results


def processStatWorkSheet(ws, results):
    stats_lst = set()
    for bmk_stats in results.values():
        stats_lst.update(list(bmk_stats.keys()))

    cols = ["Stats"] + list(results.keys())

    """ WorkSheet Format
        | Stats | Benchmark0 | Benchmark 1 | ... | Benchmark N-1 |
        | xxx   |            |             | ... |               |
        | yyy   |            |             | ... |               |
    """
    for c_idx, col_name in enumerate(cols):
        col = c_idx + 1
        _ = ws.cell(column=col, row=1, value=col_name)
        for r_idx, row_name in enumerate(stats_lst):
            row = r_idx + 2
            if col == 1:
                _ = ws.cell(column=col, row=row, value=row_name)
            else:
                if row_name in results[col_name]:
                    row_value = results[col_name][row_name]
                else:
                    row_value = "NaN"
                _ = ws.cell(column=col, row=row, value=row_value)


def save_stats(cpt_dir, result_dir_lst, out_file):
    wb = Workbook()
    for i, result_dir in enumerate(result_dir_lst):
        print(f"Processing {result_dir}...")
        results = merge_spec_results(cpt_dir, result_dir)
        title = result_dir.replace('/', '_')
        if i == 0:
            ws = wb.active
            ws.title = title
        else:
            ws = wb.create_sheet(title=title)
        processStatWorkSheet(ws, results)
    wb.save(out_file)


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('results_dirs', metavar='D',
                        type=str, nargs='+')
    parser.add_argument('--cpt-dir',
                        default="/nfs-nvme/home/share/checkpoints_profiles/spec06_rv64gcb_o2_20m")
    parser.add_argument('--out', '-o', default="tmp.xlsx")
    args = parser.parse_args()
    save_stats(args.cpt_dir, args.results_dirs, args.out)

